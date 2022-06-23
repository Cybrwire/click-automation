import openpyxl
import pyautogui as pag
import time
import webbrowser
from pynput.keyboard import Key, Controller
import re

# variables
screenSize = pag.size()
images = []
kb = Controller()
keyPresses = ['Key.enter','Key.end','Key.tab']
imgDirectory = '../images'
pag.PAUSE = 0.5
pag.FailSafeException = True

# Load excel project sheet
excelFile = ('/Users/jmontgomery/Desktop/queue-topic-project.xlsx')
wb = openpyxl.load_workbook(excelFile)
sheet = wb['Exported_Projects - Data List']


for row in sheet.iter_rows(min_row=223,max_row=260,min_col=1,max_col=2):
    cells = []
    rowNumber = int(re.search('A([0-9]{1,3})',str(row)).group(1))
    
    for cell in row:
        cells.append(cell.value)
    if cells[1] != None:
        continue
    else:
        link = sheet.cell(row=rowNumber, column=1).hyperlink.target

# open project link in browser
        webbrowser.open(link, new=2)
        time.sleep(3)

# begin process
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/showmore.png',grayscale=False,confidence=0.95)
    pag.moveTo(a/2,b/2)
    pag.click()
    pag.click()
    kb.press(Key.end)
    kb.release(Key.end)
    time.sleep(.5)

    # Delete Jess or Josh
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/routingrules.png', grayscale=False,confidence=0.8)
    pag.moveTo(a/2,(b/2)-25)
    pag.click()
    time.sleep(1)
    if not pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/routetojess.png'):
        a, b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/routetojosh.png')
    else: 
        a, b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/routetojess.png')
    pag.moveTo(a/2,b/2)
    pag.moveRel(-50,0, duration=0.1)
    pag.click()
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/delete.png')
    pag.moveTo(a/2,b/2)
    pag.click()
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/yesdeleteit.png')
    pag.moveTo(a/2,b/2)
    pag.click()
    time.sleep(2)

    # Change to Huma
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/routetoanitha.png')
    pag.moveTo(a/2,b/2)
    pag.click()
    time.sleep(2)
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/editroutingrule.png')
    pag.moveTo(a/2,b/2)
    pag.click()
    kb.type('Route to Huma')
    kb.press(Key.tab)
    kb.release(Key.tab)
    kb.press(Key.tab)
    kb.release(Key.tab)
    kb.type('Huma Khimani')
    time.sleep(2)
    kb.press(Key.enter)
    kb.release(Key.enter)
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/saverouting.png')
    pag.moveTo((a/2)-10,b/2)
    pag.click()
    time.sleep(2)
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/backtoproject.png',grayscale=False,confidence=0.9)
    pag.moveTo(a/2,(b/2)-40)
    pag.click()
    time.sleep(2)

    # move to queue topics
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/showmore.png',grayscale=False,confidence=0.9)
    pag.moveTo(a/2,b/2)
    pag.click()
    kb.press(Key.end)
    kb.release(Key.end)
    time.sleep(.5)
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/queuetopics.png')
    pag.moveTo(a/2,b/2)
    pag.click()
    time.sleep(2)

    # change strategy
    if not pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/strategy.png'):
        pag.moveRel(+115,0)
        pag.click()
        kb.press(Key.esc)
        kb.press(Key.page_down)
        a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/strategy.png')
    else: 
        a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/strategy.png')
    pag.moveTo(a/2,b/2)
    pag.click()
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/edit.png')
    pag.moveTo(a/2,b/2)
    pag.click()
    time.sleep(1.5)
    kb.press(Key.end)
    kb.release(Key.end)
    time.sleep(0.5)
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/defaultroute.png',grayscale=False,confidence=0.9)
    pag.moveTo(a/2,b/2)
    pag.click()
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/routetobrett.png')
    pag.moveTo(a/2,b/2)
    pag.click()
    pag.moveRel(-10,100)
    pag.click()
    time.sleep(1)

    # change performance
    if not pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/performance.png'):
        pag.moveRel(-10,0)
        pag.click()
        kb.press(Key.esc)
        kb.press(Key.page_down)
        time.sleep(1)
        a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/performance.png')
    else: 
        a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/performance.png')
    pag.moveTo(a/2,b/2)
    pag.click()
    a,b = pag.locateCenterOnScreen('/Users/jmontgomery/repos/click-automation/images/edit.png')
    pag.moveTo(a/2,b/2)
    pag.click()
    time.sleep(1.5)
    kb.press(Key.end)
    kb.release(Key.end)
    pag.moveTo(230,750)
    pag.click()
    pag.moveRel(0,+20)
    pag.click()
    pag.moveRel(-20,100)
    pag.click()

    time.sleep(2)
    kb.press(Key.cmd)
    kb.press('w')
    kb.release(Key.cmd)
    kb.release('w')

    sheet.cell(row=rowNumber,column=2,value='x')
    wb.save('/Users/jmontgomery/Desktop/queue-topic-project.xlsx')
    time.sleep(1)