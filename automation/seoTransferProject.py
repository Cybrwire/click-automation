# Author: Justin Montgomery
# Date: 06/20/2022
# Version: Python 3.10.5
# Purpose: This will complete a one-time task transferring projects from one SEO Specialist to another.
# I may repurpose this when a similar project comes up at work in the future. 

# native imports
import re
import time
import webbrowser
import json
import os

# external imports
import openpyxl
import pyautogui as pag
from pynput.keyboard import Controller, Key


def delete_jess_or_josh(img_key):
    """ Delete Jess or Josh"""
    a,b = pag.locateCenterOnScreen(img_key.get('routingrules'), grayscale=False, confidence=0.8)
    pag.moveTo(a/2,(b/2)-25)
    pag.click()
    time.sleep(1)

    if not pag.locateCenterOnScreen(img_key.get('routetojess')):
        a, b = pag.locateCenterOnScreen(img_key.get('routetojosh'))
    else: 
        a, b = pag.locateCenterOnScreen(img_key.get('routetojess'))
    
    pag.moveTo(a/2,b/2)
    pag.moveRel(-50,0, duration=0.1)
    pag.click()

    a,b = pag.locateCenterOnScreen(img_key.get('delete'))
    pag.moveTo(a/2,b/2)
    pag.click()

    a,b = pag.locateCenterOnScreen(img_key.get('yesdeleteit'))
    pag.moveTo(a/2,b/2)
    pag.click()
    time.sleep(2)

def change_to_huma(img_key):
    a,b = pag.locateCenterOnScreen(img_key.get('routetoanitha'))
    pag.moveTo(a/2,b/2)
    pag.click()
    time.sleep(2)

    a,b = pag.locateCenterOnScreen(img_key.get('editroutingrule'))
    pag.moveTo(a/2,b/2)
    pag.click()
    
    kb.type('Route to Huma')
    for _ in range(2):
        kb.press(TAB); kb.release(TAB)
    kb.type('Huma Khimani')
    time.sleep(2)
    kb.press(ENTER); kb.release(ENTER)

    a,b = pag.locateCenterOnScreen(img_key.get('saverouting'))
    pag.moveTo((a/2)-10,b/2)
    pag.click()
    time.sleep(2)
    a,b = pag.locateCenterOnScreen(img_key.get('backtoproject'), grayscale=False, confidence=0.9)
    pag.moveTo(a/2,(b/2)-40)
    pag.click()
    time.sleep(2)

def move_to_queue_topics(img_key):
    try: 
        a,b = pag.locateCenterOnScreen(img_key.get('showmore'), grayscale=False, confidence=0.9)
        pag.moveTo(a/2,b/2)
        pag.click()
    except TypeError:
        pag.moveTo(175,400)
        pag.click()
    kb.press(END); kb.release(END)
    time.sleep(.5)

    a,b = pag.locateCenterOnScreen(img_key.get('queuetopics'))
    pag.moveTo(a/2,b/2)
    pag.click()
    time.sleep(2)

def change_strategy(img_key):
    a, b = pag.locateCenterOnScreen(img_key.get('strategy'))
    while not a:
        pag.moveRel(+115,0)
        pag.click()
        kb.press(Key.esc)
        kb.press(Key.page_down)
        time.sleep(1)
        a, b = pag.locateCenterOnScreen(img_key.get('strategy'))
    
    pag.moveTo(a/2,b/2)
    pag.click()

    a,b = pag.locateCenterOnScreen(img_key.get('edit'))
    pag.moveTo(a/2,b/2)
    pag.click()
    time.sleep(1.5)

    kb.press(END); kb.release(END)
    time.sleep(0.5)

    a,b = pag.locateCenterOnScreen(img_key.get('defaultroute'), grayscale=False, confidence=0.9)
    pag.moveTo(a/2,b/2)
    pag.click()

    a,b = pag.locateCenterOnScreen(img_key.get('routetobrett'))
    pag.moveTo(a/2,b/2)
    pag.click()

    pag.moveRel(-10,100)
    pag.click()
    time.sleep(1)

def change_performance(img_key):
    present = pag.locateCenterOnScreen(img_key.get('performance'))
    while not present:
        pag.moveRel(-10,0)
        pag.click()
        kb.press(Key.esc)
        kb.press(Key.page_down)
        time.sleep(1)
        present = pag.locateCenterOnScreen(img_key.get('performance'))
    
    pag.moveTo(a/2,b/2)
    pag.click()

    a,b = pag.locateCenterOnScreen(img_key.get('edit'))
    pag.moveTo(a/2,b/2)
    pag.click()

    time.sleep(1.5)
    kb.press(END); kb.release(END)

    pag.moveTo(230,750)
    pag.click()
    pag.moveRel(0,+20)
    pag.click()
    pag.moveRel(-20,100)
    pag.click()

    time.sleep(2)
    kb.press(CMD)
    kb.press('w')
    kb.release(CMD)
    kb.release('w')

    sheet.cell(row=rowNumber,column=2,value='x')
    wb.save('/Users/jmontgomery/Desktop/queue-topic-project.xlsx')
    time.sleep(1)

def process(img_key):
    """Begins process"""
    a,b = pag.locateCenterOnScreen(img_key.get('showmore'), grayscale=False, confidence=0.95)
    pag.moveTo(a/2,b/2)
    for _ in range(2):
        pag.click()
    kb.press(END); kb.release(END)
    time.sleep(.5)

    delete_jess_or_josh(img_key)

    change_to_huma(img_key)

    move_to_queue_topics(img_key)

    change_strategy(img_key)

    change_performance(img_key)

if __name__ == '__main__':
    # change working directory to the directory of the script
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

    # declare variables
    screenSize = pag.size()
    images = []
    kb = Controller()
    keyPresses = ['Key.enter','Key.end','Key.tab']
    ENTER, END, ESCAPE, PAGE_DOWN, TAB, CMD = Key.enter, Key.end, Key.escape, Key.page_down, Key.tab, Key.cmd
    imgDirectory = '../images'
    pag.PAUSE = 0.5
    pag.FailSafeException = True
    # load key from images.json
    with open('images.json') as f:
        img_key = json.load(f)
    

    # Load excel project sheet
    excelFile = ('/Users/jmontgomery/Desktop/queue-topic-project.xlsx')
    wb = openpyxl.load_workbook(excelFile)
    sheet = wb['Exported_Projects - Data List']

    # do something here
    for row in sheet.iter_rows(min_row=223,min_col=1,max_col=2):
        rowNumber = int(re.search('A([0-9]{1,3})',str(row)).group(1))
        cells = [cell.value for cell in row]
        for cell in cells:
            if not cells[1]:
                link = sheet.cell(row=rowNumber,column=1).hyperlink.target
                webbrowser.open(link, new=2)
                time.sleep(3)
    
    process(img_key)